from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.db.models import Count
from .models import (
    Region,
    School,
    Application,
    BotUser,
    ApplicationStatusAudit,
    Question,
    StudentTest,
    StudentAnswer,
    AIAnswerEvaluation,
    ApplicationTestManagement,
    RegionTestControl,
)
from .tasks import analyze_application


class OverallStatusListFilter(admin.SimpleListFilter):
    title = 'umumiy_holat'
    parameter_name = 'umumiy_holat'

    def lookups(self, request, model_admin):
        return Application.OVERALL_STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(overall_status=self.value())
        return queryset


class AIHolatListFilter(admin.SimpleListFilter):
    title = 'ai_holat'
    parameter_name = 'ai_holat'

    def lookups(self, request, model_admin):
        return StudentTest.AI_HOLAT_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(student_tests__ai_holat=self.value())
        return queryset


class AISifatBahosiListFilter(admin.SimpleListFilter):
    title = 'ai_sifat_bahosi'
    parameter_name = 'ai_sifat_bahosi'

    def lookups(self, request, model_admin):
        return StudentTest.LEVEL_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(student_tests__ai_sifat_bahosi=self.value())
        return queryset


@admin.register(Region)
class RegionAdmin(admin.ModelAdmin):
    list_display = ['name', 'is_open', 'deadline']
    list_filter = ['is_open']
    search_fields = ['name']
    list_editable = ['is_open']


@admin.register(RegionTestControl)
class RegionTestControlAdmin(admin.ModelAdmin):
    list_display = ('region', 'is_test_active', 'updated_at')
    list_filter = ('is_test_active', 'region')
    search_fields = ('region__name',)
    list_editable = ('is_test_active',)


def export_school_stats_xls(modeladmin, request, queryset):
    """Tanlangan maktablar bo'yicha statistikani Excelga yuklash"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Maktab Statistikasi"
    
    # Headers
    headers = ["Hudud", "Maktab", "Arizalar soni"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Annotate queryset with application count
    # We need to import Application here or use related_name if available.
    # School has related_name='schools' from Region, but for Application it is 'application_set' by default directly,
    # or actually in the Application model: school = models.ForeignKey(..., related_name='applications')?
    # Checking Application model: school = models.ForeignKey(..., related_name='applications' is NOT set, so default is application_set)
    # Actually, let's check models.py content again to be sure about related_name if any.
    # Application model: school = models.ForeignKey(School, ..., related_name='application_set' (default))
    # Wait, looking at previous file view of models.py (Step 27):
    # school = models.ForeignKey(School, ... verbose_name="Maktab")
    # So related name is default 'application_set'
    
    schools_with_counts = queryset.annotate(
        app_count=Count('application')
    ).order_by('region__name', 'name')
    
    # Write data
    for row_num, school in enumerate(schools_with_counts, 2):
        ws.cell(row=row_num, column=1, value=school.region.name)
        ws.cell(row=row_num, column=2, value=school.name)
        ws.cell(row=row_num, column=3, value=school.app_count)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=maktab_statistikasi.xlsx'
    wb.save(response)
    return response

export_school_stats_xls.short_description = "Statistikani Excelga yuklash"


@admin.register(School)
class SchoolAdmin(admin.ModelAdmin):
    list_display = ['name', 'region', 'get_app_count']
    list_filter = ['region']
    search_fields = ['name']
    autocomplete_fields = ['region']
    actions = [export_school_stats_xls]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(app_count=Count('application'))
    
    def get_app_count(self, obj):
        return obj.app_count
    get_app_count.short_description = "Arizalar soni"
    get_app_count.admin_order_field = 'app_count'


def export_to_excel(modeladmin, request, queryset):
    """Excel faylga yuklab olish"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"
    
    # Headers
    headers = [
        "ID",
        "To'liq ism",
        "Telefon",
        "Hudud",
        "Maktab",
        "Sinf",
        "Jihozingiz",
        "Ingliz tili",
        "O'zingiz haqingizda",
        "Holat",
        "AI Holati",
        "AI Bahosi",
        "AI Izohi",
        "Yaratilgan sana",
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, app in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=app.id)
        ws.cell(row=row_num, column=2, value=app.full_name)
        ws.cell(row=row_num, column=3, value=app.phone)
        ws.cell(row=row_num, column=4, value=app.region.name if app.region else "-")
        ws.cell(row=row_num, column=5, value=app.school.name if app.school else "-")
        ws.cell(row=row_num, column=6, value=app.get_grade_display())
        ws.cell(row=row_num, column=7, value=app.get_device_display())
        ws.cell(row=row_num, column=8, value=app.get_english_level_display())
        ws.cell(row=row_num, column=9, value=app.about)
        ws.cell(row=row_num, column=10, value=app.get_status_display())
        ws.cell(row=row_num, column=11, value=app.get_ai_status_display())
        ws.cell(row=row_num, column=12, value=app.description_quality or "-")
        ws.cell(row=row_num, column=13, value=app.ai_reason or "-")
        ws.cell(row=row_num, column=14, value=app.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a new sheet for statistics
    stats_ws = wb.create_sheet(title="Statistika")
    
    # Headers for stats
    stats_headers = ["Hudud", "Maktab", "Ishtirokchilar soni"]
    for col_num, header in enumerate(stats_headers, 1):
        cell = stats_ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Aggregate statistics from the queryset
    # We group by region name and school name and count
    stats_data = queryset.values(
        'region__name', 
        'school__name'
    ).annotate(
        count=Count('id')
    ).order_by('region__name', 'school__name')
    
    # Write stats data
    for row_num, entry in enumerate(stats_data, 2):
        stats_ws.cell(row=row_num, column=1, value=entry['region__name'] or "-")
        stats_ws.cell(row=row_num, column=2, value=entry['school__name'] or "-")
        stats_ws.cell(row=row_num, column=3, value=entry['count'])
    
    # Auto-adjust column widths for stats sheet
    for column in stats_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        stats_ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=arizalar.xlsx'
    wb.save(response)
    return response

export_to_excel.short_description = "Excel faylga yuklab olish"


def export_to_json(modeladmin, request, queryset):
    """JSON faylga yuklab olish"""
    import json
    from django.http import HttpResponse # Ensure output is correct type
    
    data = []
    for app in queryset:
        data.append({
            "id": app.id,
            "full_name": app.full_name,
            "phone": app.phone,
            "region": app.region.name if app.region else None,
            "school": app.school.name if app.school else None,
            "grade": app.get_grade_display(),
            "device": app.get_device_display(),
            "english_level": app.get_english_level_display(),
            "about": app.about,
            "status": app.get_status_display(),
            "ai_status": app.get_ai_status_display(),
            "ai_reason": app.ai_reason,
            "description_quality": app.description_quality,
            "analyzed_at": app.analyzed_at.strftime('%Y-%m-%d %H:%M') if app.analyzed_at else None,
            "created_at": app.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename=arizalar.json'
    return response

export_to_json.short_description = "JSON faylga yuklab olish"


def bulk_set_status_approved(modeladmin, request, queryset):
    """Bulk action: Set status to approved with audit logging"""
    from django.db import transaction
    
    updated_count = 0
    
    try:
        with transaction.atomic():
            for app in queryset:
                previous_status = app.status
                if previous_status != 'accepted':
                    app.status = 'accepted'
                    app.save()
                    
                    # Create audit log
                    ApplicationStatusAudit.objects.create(
                        application=app,
                        previous_status=previous_status,
                        new_status='accepted',
                        admin_user=request.user,
                        action_type='bulk_status_update'
                    )
                    updated_count += 1
        
        modeladmin.message_user(request, f"{updated_count} ta ariza QABUL QILINDI holatiga o'zgartirildi.")
    except Exception as e:
        modeladmin.message_user(request, f"Xato: {str(e)}", level='ERROR')

bulk_set_status_approved.short_description = "Tanlanganlarga holat: QABUL QILINDI"


def bulk_set_status_rejected(modeladmin, request, queryset):
    """Bulk action: Set status to rejected with audit logging"""
    from django.db import transaction
    
    updated_count = 0
    
    try:
        with transaction.atomic():
            for app in queryset:
                previous_status = app.status
                if previous_status != 'rejected':
                    app.status = 'rejected'
                    app.save()
                    
                    # Create audit log
                    ApplicationStatusAudit.objects.create(
                        application=app,
                        previous_status=previous_status,
                        new_status='rejected',
                        admin_user=request.user,
                        action_type='bulk_status_update'
                    )
                    updated_count += 1
        
        modeladmin.message_user(request, f"{updated_count} ta ariza RAD ETILDI holatiga o'zgartirildi.")
    except Exception as e:
        modeladmin.message_user(request, f"Xato: {str(e)}", level='ERROR')

bulk_set_status_rejected.short_description = "Tanlanganlarga holat: RAD ETILDI"


def bulk_set_status_pending(modeladmin, request, queryset):
    """Bulk action: Set status to pending with audit logging"""
    from django.db import transaction
    
    updated_count = 0
    
    try:
        with transaction.atomic():
            for app in queryset:
                previous_status = app.status
                if previous_status != 'pending':
                    app.status = 'pending'
                    app.save()
                    
                    # Create audit log
                    ApplicationStatusAudit.objects.create(
                        application=app,
                        previous_status=previous_status,
                        new_status='pending',
                        admin_user=request.user,
                        action_type='bulk_status_update'
                    )
                    updated_count += 1
        
        modeladmin.message_user(request, f"{updated_count} ta ariza KUTILMOQDA holatiga o'zgartirildi.")
    except Exception as e:
        modeladmin.message_user(request, f"Xato: {str(e)}", level='ERROR')

bulk_set_status_pending.short_description = "Tanlanganlarga holat: KUTILMOQDA"


@admin.register(Application)
class ApplicationAdmin(admin.ModelAdmin):
    list_display = [
        'id', 
        'full_name', 
        'phone', 
        'region', 
        'school', 
        'status',
        'ai_status',
        'description_quality',
        'created_at'
    ]
    list_filter = ['status', 'ai_status', 'description_quality', 'region', 'grade', 'english_level', 'device', 'created_at']
    search_fields = ['full_name', 'phone']
    readonly_fields = ['created_at', 'updated_at', 'analyzed_at']
    list_editable = ['status']
    autocomplete_fields = ['region', 'school']
    
    def run_ai_analysis(self, request, queryset):
        """AI tahlilini qo'lda ishga tushirish"""
        triggered_count = 0
        skipped_count = 0
        for app in queryset:
            if app.about and len(app.about.strip()) > 10:  # Minimum length check
                analyze_application.delay(app.id)
                triggered_count += 1
            else:
                skipped_count += 1
        
        msg = f"{triggered_count} ta ariza AI tahliliga yuborildi."
        if skipped_count > 0:
            msg += f" {skipped_count} ta ariza ma'lumot yetarli emasligi sababli o'tkazib yuborildi."
        
        self.message_user(request, msg)
    
    run_ai_analysis.short_description = "AI tahlilini ishga tushirish"
    
    actions = [
        export_to_excel, 
        export_to_json, 
        run_ai_analysis,
        bulk_set_status_approved,
        bulk_set_status_rejected,
        bulk_set_status_pending
    ]
    
    fieldsets = (
        ('Shaxsiy ma\'lumotlar', {
            'fields': ('full_name', 'phone')
        }),
        ('Ta\'lim ma\'lumotlari', {
            'fields': ('region', 'school', 'grade')
        }),
        ('Qo\'shimcha ma\'lumotlar', {
            'fields': ('about', 'device', 'english_level')
        }),
        ('AI Tahlil', {
            'fields': (
                'ai_status', 
                'description_quality',
                'ai_reason', 
                'analyzed_at'
            )
        }),
        ('Holat', {
            'fields': ('status',)
        }),
        ('Vaqt belgilari', {
            'fields': ('created_at', 'updated_at'),
        }),
    )


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('id', 'short_prompt', 'is_active', 'created_at', 'updated_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('prompt',)
    list_editable = ('is_active',)

    def short_prompt(self, obj):
        if len(obj.prompt) <= 100:
            return obj.prompt
        return f"{obj.prompt[:100]}..."
    short_prompt.short_description = "Savol"


class AIAnswerEvaluationInline(admin.StackedInline):
    model = AIAnswerEvaluation
    extra = 0
    can_delete = False
    fk_name = 'answer'
    readonly_fields = ('created_at', 'updated_at')


class StudentAnswerInline(admin.StackedInline):
    model = StudentAnswer
    extra = 0
    readonly_fields = ('question', 'written_answer', 'created_at', 'updated_at')
    can_delete = False
    show_change_link = True


@admin.register(StudentAnswer)
class StudentAnswerAdmin(admin.ModelAdmin):
    list_display = ('id', 'test', 'question', 'preview_answer', 'get_score_level')
    list_select_related = ('test__student', 'question')
    search_fields = ('test__student__full_name', 'test__student__phone', 'question__prompt', 'written_answer')
    inlines = [AIAnswerEvaluationInline]

    def preview_answer(self, obj):
        text = obj.written_answer or ''
        if len(text) <= 80:
            return text
        return f"{text[:80]}..."
    preview_answer.short_description = "Javob"

    def get_score_level(self, obj):
        return getattr(getattr(obj, 'ai_evaluation', None), 'score_level', '-')
    get_score_level.short_description = "AI Daraja"


@admin.register(StudentTest)
class StudentTestAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'student_full_name',
        'student_phone',
        'is_submitted',
        'ai_sifat_bahosi',
        'ai_holat',
        'submitted_at',
    )
    list_filter = ('is_submitted', 'ai_holat', 'ai_sifat_bahosi', 'submitted_at')
    search_fields = ('student__full_name', 'student__phone')
    list_select_related = ('student', 'student__region', 'student__school')
    readonly_fields = ('created_at', 'updated_at', 'submitted_at')
    inlines = [StudentAnswerInline]

    def student_full_name(self, obj):
        return obj.student.full_name
    student_full_name.short_description = 'Full name'

    def student_phone(self, obj):
        return obj.student.phone
    student_phone.short_description = 'Phone'


@admin.register(ApplicationTestManagement)
class ApplicationTestManagementAdmin(admin.ModelAdmin):
    list_display = (
        'full_name',
        'phone',
        'region',
        'school',
        'overall_status',
        'ai_holat',
        'ai_sifat_bahosi',
        'overall_ai_summary',
        'test_submitted',
    )
    list_filter = ('region', OverallStatusListFilter, AIHolatListFilter, AISifatBahosiListFilter)
    search_fields = ('full_name', 'phone', 'region__name', 'school__name')
    ordering = ('-created_at',)
    list_editable = ('overall_status',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('region', 'school').prefetch_related('student_tests')

    def _test_obj(self, obj):
        return obj.student_tests.first()

    def ai_sifat_bahosi(self, obj):
        test = self._test_obj(obj)
        return test.ai_sifat_bahosi if test else '-'
    ai_sifat_bahosi.short_description = 'ai_sifat_bahosi'

    def ai_holat(self, obj):
        test = self._test_obj(obj)
        return test.ai_holat if test else StudentTest.AI_HOLAT_KUTILAYAPTI
    ai_holat.short_description = 'ai_holat'

    def overall_ai_summary(self, obj):
        test = self._test_obj(obj)
        if not test or not test.overall_ai_summary:
            return '-'
        return test.overall_ai_summary[:140]
    overall_ai_summary.short_description = 'overall_ai_summary'

    def test_submitted(self, obj):
        test = self._test_obj(obj)
        return bool(test and test.is_submitted)
    test_submitted.short_description = 'test_submitted'
    test_submitted.boolean = True


@admin.register(BotUser)
class BotUserAdmin(admin.ModelAdmin):
    # Admin panelda ko‘rinadigan ustunlar
    list_display = (
        'telegram_id',
        'claimed_phone',
        'created_at',
    )

    # Qidiruv
    search_fields = (
        'telegram_id',
        'claimed_phone',
    )

    # Filtrlash
    list_filter = (
        'created_at',
    )

    # Tartiblash (eng yangisi yuqorida)
    ordering = ('-created_at',)

    # O‘qish rejimi (o‘zgartirilmasin)
    readonly_fields = (
        'telegram_id',
        'created_at',
    )

    # Sahifalash
    list_per_page = 25

    # Form layout
    fieldsets = (
        ('Telegram maʼlumotlari', {
            'fields': ('telegram_id',),
        }),
        ('Bog‘langan telefon', {
            'fields': ('claimed_phone',),
        }),
        ('Vaqt maʼlumotlari', {
            'fields': ('created_at',),
        }),
    )

@admin.register(ApplicationStatusAudit)
class ApplicationStatusAuditAdmin(admin.ModelAdmin):
    list_display = (
        'application',
        'previous_status',
        'new_status',
        'admin_user',
        'action_type',
        'timestamp'
    )
    
    list_filter = (
        'action_type',
        'previous_status',
        'new_status',
        'timestamp'
    )
    
    search_fields = (
        'application__full_name',
        'application__phone',
        'admin_user__username'
    )
    
    readonly_fields = (
        'application',
        'previous_status',
        'new_status',
        'admin_user',
        'action_type',
        'timestamp'
    )
    
    ordering = ('-timestamp',)
    
    # Prevent any modifications - immutable audit log
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
