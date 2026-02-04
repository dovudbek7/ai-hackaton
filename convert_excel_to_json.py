import openpyxl
import json
import os

def convert_excel_to_json(excel_file, sheet_name, output_json):
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found.")
        return

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in {excel_file}.")
            print(f"Available sheets: {wb.sheetnames}")
            return

        sheet = wb[sheet_name]
        
        # Get headers from the first row
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print("Error: Sheet is empty.")
            return

        headers = rows[0]
        data = []

        for row in rows[1:]:
            # Map headers to row values
            entry = {}
            for header, value in zip(headers, row):
                if header:
                    entry[header] = value
            if any(entry.values()):  # Only add non-empty rows
                data.append(entry)

        # Write to JSON
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        print(f"Successfully converted {len(data)} rows to {output_json}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    excel_file = 'maktab soni.xlsx'
    sheet_name = 'maktab kesimida'
    output_json = 'maktab_soni.json'
    convert_excel_to_json(excel_file, sheet_name, output_json)
